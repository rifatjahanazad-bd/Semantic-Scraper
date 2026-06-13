# ============================================
# IMPORTS
# ============================================

from playwright.sync_api import sync_playwright
import requests
import json
import os
import time

from openpyxl import load_workbook
from openpyxl import Workbook


# ============================================
# INPUT EXCEL FILE
# ============================================

input_file = "input_list.xlsx"


# ============================================
# READ INPUTS FROM EXCEL
# ============================================

def read_inputs_from_excel(file_name="input_list.xlsx", max_rows=5):

    """
    Column A = URL
    Column B = Doctor Name
    """

    if not os.path.exists(file_name):

        print(f"\nExcel file not found: {file_name}\n")

        return []

    try:

        wb = load_workbook(
            file_name,
            read_only=True,
            data_only=True
        )

        ws = wb.active

        inputs = []

        for row in ws.iter_rows(
            min_row=2,
            max_row=1 + max_rows,
            values_only=True
        ):

            url = None
            name = None

            # Column A
            if row and len(row) >= 1 and row[0]:

                url = str(row[0]).strip()

            # Column B
            if row and len(row) >= 2 and row[1]:

                name = str(row[1]).strip()

            # Skip empty rows
            if not url:

                continue

            inputs.append({
                "url": url,
                "name": name
            })

        return inputs

    except Exception as e:

        print("\nError reading Excel:\n", e)

        return []


# ============================================
# WRITE RESULTS TO EXCEL
# ============================================

def write_results_to_excel(results, output_file="output_list.xlsx"):

    try:

        wb = Workbook()

        ws = wb.active

        ws.append([
            "URL",
            "Doctor Name",
            "Professional Role",
            "Department Name"
        ])

        for r in results:

            ws.append([
                r.get("url"),
                r.get("name"),
                r.get("professional_role"),
                r.get("department_name")
            ])

        wb.save(output_file)

        print(f"\nSaved results to: {output_file}\n")

    except Exception as e:

        print("\nFailed writing output Excel:\n", e)


# ============================================
# GET PAGE TEXT
# ============================================

def get_page_text(page, url):

    """
    Reuses SAME browser page
    """

    try:

        print(f"\nOpening: {url}\n")

        page.goto(
            url,
            timeout=90000
        )

        # Wait for page load
        page.wait_for_timeout(5000)

        print("\nExtracting visible text...\n")

        text = page.locator("body").inner_text()

        return text

    except Exception as e:

        print("\nBrowser extraction error:\n", e)

        return None


# ============================================
# AI EXTRACTION
# ============================================

def extract_hcp_info(page_text, hcp_name=""):

    prompt = f"""
You are extracting healthcare professional information.

TARGET DOCTOR:
{hcp_name}

Extract information ONLY for this doctor.

Ignore unrelated doctors or staff.

Return JSON only.

Fields:
- doctor_name
- professional_role
- department_name

Rules:
- If information is missing return null
- Do not guess
- Do not infer
- Return valid JSON only
- No markdown
- No explanations

Example format:

{{
    "doctor_name": "",
    "professional_role": "",
    "department_name": ""
}}

PAGE TEXT:
{page_text[:12000]}
"""

    try:

        print("\nSending text to local AI...\n")

        response = requests.post(

            "http://localhost:11434/api/generate",

            json={

                "model": "qwen2.5:7b",

                "prompt": prompt,

                "stream": False
            }
        )

        result = response.json()

        raw_output = result.get("response")

        return raw_output

    except Exception as e:

        print("\nAI extraction failed:\n", e)

        return None


# ============================================
# MAIN WORKFLOW
# ============================================

print("\n========== STARTING EXTRACTION ==========\n")

results = []

# Read Excel
entries = read_inputs_from_excel(
    file_name=input_file,
    max_rows=3
)

# Stop if Excel empty
if not entries:

    print("\nNo valid Excel entries found.\n")

    exit()

# ============================================
# OPEN BROWSER ONLY ONCE
# ============================================

with sync_playwright() as p:

    browser = p.chromium.launch(
        headless=False
    )

    # Create ONE browser tab
    page = browser.new_page()

    # ========================================
    # PROCESS ALL DOCTORS
    # ========================================

    for idx, entry in enumerate(entries, start=1):

        name = entry.get("name")

        url = entry.get("url")

        print(f"\n========== ENTRY {idx} ==========\n")

        print(f"Doctor: {name}")

        print(f"URL: {url}\n")

        # ====================================
        # EXTRACT PAGE TEXT
        # ====================================

        page_text = get_page_text(page, url)

        if not page_text:

            print("\nFailed extracting webpage text.\n")

            continue

        # ====================================
        # AI EXTRACTION
        # ====================================

        result = extract_hcp_info(
            page_text,
            hcp_name=name
        )

        if not result:

            print("\nAI extraction failed.\n")

            continue

        print("\n========== RAW MODEL OUTPUT ==========\n")

        print(result)

        # ====================================
        # CLEAN AI OUTPUT
        # ====================================

        cleaned = (
            result
            .replace("```json", "")
            .replace("```", "")
            .strip()
        )

        # Default values
        doctor_name = None
        professional_role = None
        department_name = None

        # ====================================
        # PARSE JSON
        # ====================================

        try:

            parsed = json.loads(cleaned)

            if isinstance(parsed, dict):

                doctor_name = parsed.get(
                    "doctor_name"
                )

                professional_role = parsed.get(
                    "professional_role"
                )

                department_name = parsed.get(
                    "department_name"
                )

        except Exception as e:

            print("\nJSON parsing failed:\n", e)

        # ====================================
        # SAVE RESULTS
        # ====================================

        results.append({

            "url": url,

            "name": doctor_name or name,

            "professional_role": professional_role,

            "department_name": department_name
        })

        print("\n========== ENTRY FINISHED ==========\n")

        time.sleep(1)

    # ========================================
    # CLOSE BROWSER AFTER ALL DOCTORS
    # ========================================

    browser.close()


# ============================================
# WRITE OUTPUT EXCEL
# ============================================

print("\n========== WRITING OUTPUT ==========\n")

write_results_to_excel(
    results,
    output_file="output_list.xlsx"
)

print("\n========== ALL TASKS FINISHED ==========\n")